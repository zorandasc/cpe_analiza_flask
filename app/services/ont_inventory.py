import calendar
from datetime import datetime
from sqlalchemy import text
from app.extensions import db
from app.utils.dates import get_previous_month_end
from openpyxl import load_workbook
from app.utils.permissions import ftth_view_required
from app.queries.ont_onventory import (
    get_last_4_months,
    get_ont_inventory_pivoted,
)


TOTAL_KEY = "__TOTAL__"


def get_ont_records_view_data():
    # calculate current month date
    previous_month_end = get_previous_month_end()

    # label → presentation used only in the table header
    # key → internal identifier (used in SQL + structure of html table)
    # w.isoformat()='YYYY-MM-DD'
    months = [
        {"key": m.isoformat(), "label": m.strftime("%d-%m-%Y")}
        for m in sorted(get_last_4_months())
    ]

    month_keys = [m["key"] for m in months]

    # get the pivoted data from db
    records = get_ont_inventory_pivoted(month_keys)

    records_grouped = _group_records(records, month_keys)

    return {
        "previous_month_end": previous_month_end,
        "months": months,
        "records": records_grouped,
    }


def update_recent_ont_inventory(form_data):
    if not ftth_view_required():  # AUTHORIZATION
        return False, "Niste autorizovani."

    previous_month_end = get_previous_month_end()

    try:
        for key, value in form_data.items():
            if key == "__TOTAL__":
                continue
            try:
                city_id = int(key)
                quantity = int(value or 0)

            except ValueError:
                # Skip this record if ID or Quantity is invalid
                continue

            # because of UNIQUE (city_id, week_end) constraints
            # added when defining table OntInventory in postgres
            # business logic is: “For this month, insert if missing, update if exists”
            # That is exactly what PostgreSQL ON CONFLICT DO UPDATE is for.
            # ORM add_all() cannot do UPSERT cleanly
            db.session.execute(
                text("""
                    INSERT INTO ont_inventory (city_id, month_end, quantity)
                    VALUES (:city_id, :month_end, :quantity)
                    ON CONFLICT (city_id, month_end)
                    DO UPDATE SET quantity = EXCLUDED.quantity, updated_at = NOW();
                """),
                {
                    "city_id": city_id,
                    "month_end": previous_month_end,
                    "quantity": quantity,
                },
            )

        # PostgreSQL loves batching UPSERTs in a single transaction.
        # UPSERT = “UPDATE or INSERT”.
        # Only on commit() does SQLAlchemy:
        # Send all pending SQL statements to the database
        # Wrap them in a transaction
        # `Make them permanent in the DB
        db.session.commit()
        return True, f"Novo stanje za {previous_month_end} uspješno sačuvano!"

    except Exception as e:
        db.session.rollback()
        print(e)
        return False, "Greška prilikom čuvanja podataka."


def get_ont_records_excel_export():
    previous_month_end = get_previous_month_end()

    # month is used in SQL + and for structure of html table
    # covert datetime.date to date string
    # label → presentation used only in the table header
    # key → internal identifier (used in SQL + structure of html table)
    # w.isoformat()='YYYY-MM-DD'
    months = [
        {"key": m.isoformat(), "label": m.strftime("%d-%m-%Y")}
        for m in sorted(get_last_4_months())
    ]

    month_keys = [m["key"] for m in months]

    # get the pivoted data from db
    records = get_ont_inventory_pivoted(month_keys)

    records_grouped = _group_records(records, month_keys)

    # ---- HEADERS ----
    # ----STB HEADERS ----
    headers = ["Skladišta"] + month_keys

    rows = []
    for ont in records_grouped:
        row = [ont["name"]] + [ont["dates"][m]["quantity"] for m in month_keys]
        rows.append(row)

    return headers, rows, previous_month_end


def parce_excel_segments(file_storage):
    """
    Parse ecxel row segment by segment. The border of segmennts are 'medium' style.
    Distinguish between top and bottom border style.
    End of loop is both top and bottom border style'medium' on last row in excel

    """
    file_storage.stream.seek(0)

    # Load the workbook directly from the file stream
    try:
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        return f"Error loading workbook: {e}"

    sheet = wb.active  # Or wb['Sheet1']
    # list of all segments
    summaries = []
    current_subtotal = 0
    # Assuming header is row 1, data starts at row 2
    starting_row = 2
    # Column O in excel is the 15th column
    column_index = 15

    for row in sheet.iter_rows(
        min_row=starting_row, max_col=column_index, min_col=column_index
    ):
        cell = row[0]
        border = cell.border
        val = cell.value if isinstance(cell.value, (int, float)) else 0

        # 1. STOP & CAPTURE: Grand Total Detection
        if border.top.style == "medium" and border.bottom.style == "medium":
            grand_total_from_file = val
            print(f"--- VALIDATION: Grand Total Found: {grand_total_from_file} ---")
            break

        # 2. TRIGGER TOP BORDER: New Group Starts (Top Style)
        if border.top.style == "medium":
            if current_subtotal > 0:
                summaries.append(current_subtotal)
            current_subtotal = val

        # 3. TRIGGER BOTTOM BORDER: Current Group Ends (Bottom Style)
        elif border.bottom.style == "medium":
            current_subtotal += val
            if current_subtotal > 0:
                summaries.append(current_subtotal)
            current_subtotal = 0

        # 4. NORMAL: Regular accumulation
        else:
            current_subtotal += val

    # Final cleanup if data remains (in case Grand Total row was missing)
    if current_subtotal > 0:
        summaries.append(current_subtotal)

    # --- Verification Logic ---
    calculated_total = sum(summaries)
    is_valid = calculated_total == grand_total_from_file

    return {
        "segments": summaries,
        "calculated_total": calculated_total,
        "excel_grand_total": grand_total_from_file,
        "match": is_valid,
    }


def save_imported_segments_to_db(segments, target_date=None):
    """
    Map the Excel segment indexses (0, 1, 2...) to your DB city_id
    """

    if not ftth_view_required():
        return False, "Niste autorizovani."

    # Backend sanitization
    # If target_date is provided (Admin), use it.
    # Otherwise (User), get current month end.
    if target_date:
        # If target_date is a string from JSON, convert to object
        if isinstance(target_date, str):
            dt = datetime.strptime(target_date, "%Y-%m-%d")
        else:
            dt = target_date

        # FORCE the date to the last day of that specific month
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        month_to_save = dt.replace(day=last_day).date()

    else:
        # Regular user logic: get actual current month end
        month_to_save = get_previous_month_end()

    # SEG 1 BANJA LUKA -> (city_id=3)
    # SEG 2 BIJELJINA-> (city_id=6)
    # SEG 3 BRCKO-> (city_id=7)
    # SEG 4 DOBOJ-> (city_id=5)
    # SEG 5 SARAJEVO-> (city_id=9)
    # SEG 6 PRIJEDOR-> (city_id=4)
    # SEG 7 TREBINJE-> (city_id=11)
    # SEG 8 FOCA-> (city_id=10)
    # SEG 9 ZVORNIK-> (city_id=8)
    # Map the Excel segment index (0, 1, 2...) to your DB city_id
    # Ensure this order matches your Excel segments exactly!
    city_mapping = [3, 6, 7, 5, 9, 4, 11, 10, 8]

    try:
        for index, value in enumerate(segments):
            # Guard against Excel providing more segments than we have city mappings
            if index >= len(city_mapping):
                break

            city_id = city_mapping[index]
            quantity = int(value or 0)

            # UPSERT
            db.session.execute(
                text("""
                    INSERT INTO ont_inventory (city_id, month_end, quantity)
                    VALUES (:city_id, :month_end, :quantity)
                    ON CONFLICT (city_id, month_end)
                    DO UPDATE SET quantity = EXCLUDED.quantity, updated_at = NOW();
                """),
                {
                    "city_id": city_id,
                    "month_end": month_to_save,
                    "quantity": quantity,
                },
            )

        db.session.commit()
        return (
            True,
            f"Uspješno ažurirano {len(segments)} skladišta za {month_to_save}.",
        )

    except Exception as e:
        db.session.rollback()
        print(e)
        return False, "Greška prilikom čuvanja podataka."


# -------------------------
# INTERNAL HELPERS
# -------------------------
def _group_records(records, month_keys):
    grouped = {}

    for row in records:
        cityid = row["id"] or TOTAL_KEY

        if cityid not in grouped:
            grouped[cityid] = {
                "id": cityid,
                "name": row["name"],
                "last_updated": row["last_updated"],
                "is_total": cityid == TOTAL_KEY,
                "dates": {month: {"quantity": 0} for month in month_keys},
            }

        for month in month_keys:
            grouped[cityid]["dates"][month]["quantity"] = row.get(month, 0)

    return list(grouped.values())
