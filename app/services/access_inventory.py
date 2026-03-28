import calendar
from datetime import datetime
from sqlalchemy import text
from app.extensions import db
from app.models import AccessTypes
from app.services.charts import get_visible_cities
from app.services.user_activity_log import log_user_action
from app.utils.dates import get_previous_month_end
from openpyxl import load_workbook
from app.utils.permissions import ftth_view_required
from app.queries.access_inventory import (
    get_last_4_months,
    get_access_inventory_pivoted,
    get_access_inventory_history,
)


TOTAL_KEY = "__TOTAL__"


def get_access_records_view_data():
    # calculate current month date
    previous_month_end = get_previous_month_end()

    # label → presentation used only in the table header
    # key → internal identifier (used in SQL + structure of html table)
    # w.isoformat()='YYYY-MM-DD'
    months = [
        {"key": m.isoformat(), "label": m.strftime("%d-%m-%Y")}
        for m in sorted(get_last_4_months())
    ]

    month_keys = [m["key"] for m in months]

    access_types = (
        AccessTypes.query.filter_by(is_active=True).order_by(AccessTypes.id).all()
    )

    grouped_data = {}

    # Right now you're running N queries: 1 per access_type
    # This is perfectly fine for small numbers (3–6 types).
    # But if one day: 20 access type
    # then Single query grouped by (city_id, access_type_id).
    for at in access_types:
        # get the pivoted data from db for every access_type_id
        records = get_access_inventory_pivoted(months=month_keys, access_type_id=at.id)

        # last_updated per access_type_id
        last_updated = max(
            (r["last_updated"] for r in records if r["last_updated"] is not None),
            default=None,
        )

        # group data for view
        grouped_data[at.name] = {
            "rows": _group_records(records, month_keys),
            "last_updated": last_updated,
        }

    return {
        "previous_month_end": previous_month_end,
        "months": months,
        "access_types": access_types,
        "records": grouped_data,
    }


def update_recent_access_inventory(form_data):
    if not ftth_view_required():  # AUTHORIZATION
        return False, "Niste autorizovani."

    previous_month_end = get_previous_month_end()

    try:
        access_type_id = int(form_data["access_type_id"])
    except (KeyError, ValueError):
        return False, "Nevažeći tip pristupa."

    # Security + data integrity.
    access_type = AccessTypes.query.get(access_type_id)
    if not access_type:
        return False, "Tip pristupa ne postoji."

    SQL = text("""
        INSERT INTO access_inventory (city_id, access_type_id, month_end, quantity)
        VALUES (:city_id, :access_type_id, :month_end, :quantity)
        ON CONFLICT (city_id, access_type_id, month_end)
        DO UPDATE SET quantity = EXCLUDED.quantity,
                    updated_at = NOW();
    """)

    try:
        for key, value in form_data.items():
            if key in ("__TOTAL__", "access_type_id"):
                continue
            try:
                city_id = int(key)
                quantity = int(value or 0)
            except ValueError:
                # Skip this record if ID or Quantity is invalid
                continue

            # because of UNIQUE (city_id,access_type_id, week_end) constraints
            # added when defining table AccessInventory in postgres
            # business logic is: “For this month, insert if missing, update if exists”
            # That is exactly what PostgreSQL ON CONFLICT DO UPDATE is for.
            # ORM add_all() cannot do UPSERT cleanly
            db.session.execute(
                SQL,
                {
                    "city_id": city_id,
                    "access_type_id": access_type_id,
                    "month_end": previous_month_end,
                    "quantity": quantity,
                },
            )

        log_user_action(
            action="update",
            table_name="access_inventory",
            record_id=access_type_id,
            details={
                "count": len(form_data),
                "month_end": str(previous_month_end),
                "type": access_type.name,
            },
        )

        # PostgreSQL loves batching UPSERTs in a single transaction.
        # UPSERT = “UPDATE or INSERT”.
        # Only on commit() does SQLAlchemy:
        # Send all pending SQL statements to the database
        # Wrap them in a transaction
        # `Make them permanent in the DB
        db.session.commit()
        return True, f"Novo stanje za {previous_month_end} uspješno sačuvano!"

    except Exception as e:
        db.session.rollback()
        print(e)
        return False, "Greška prilikom čuvanja podataka."


def get_access_records_history(
    access_type_id: int,
    page: int,
    per_page: int,
):
    access_type = AccessTypes.query.get(access_type_id)

    if not access_type:
        return None, None, None, "Pristupna ne postoji."

    # get_visible_cities() imported from charts services
    # List of cities visible in access_ivnetory table
    schema_list = get_visible_cities("access_inventory")

    records = get_access_inventory_history(
        access_type_id=access_type.id,
        schema_list=schema_list,
        page=page,
        per_page=per_page,
    )

    return access_type, records, schema_list, None


def get_access_records_excel_export(id):
    previous_month_end = get_previous_month_end()

    try:
        access_type_id = int(id)
    except (KeyError, ValueError):
        return False, "Nevažeći tip pristupa."

    # Security + data integrity.
    access_type = AccessTypes.query.get(access_type_id)
    if not access_type:
        return False, "Tip pristupa ne postoji."

    # month is used in SQL + and for structure of html table
    # covert datetime.date to date string
    # label → presentation used only in the table header
    # key → internal identifier (used in SQL + structure of html table)
    # w.isoformat()='YYYY-MM-DD'
    months = [
        {"key": m.isoformat(), "label": m.strftime("%d-%m-%Y")}
        for m in sorted(get_last_4_months())
    ]

    month_keys = [m["key"] for m in months]

    # get the pivoted data from db
    records = get_access_inventory_pivoted(
        months=month_keys, access_type_id=access_type_id
    )

    records_grouped = _group_records(records, month_keys)

    # ---- HEADERS ----
    # ----STB HEADERS ----
    headers = ["Skladišta"] + month_keys

    rows = []
    for access in records_grouped:
        row = [access["name"]] + [access["dates"][m]["quantity"] for m in month_keys]
        rows.append(row)

    return access_type.name, headers, rows, previous_month_end


def parce_excel_segments(file_storage):
    """
    Parse ecxel row segment by segment. The border of segmennts are 'medium' style.
    Distinguish between top and bottom border style.
    End of loop is both top and bottom border style'medium' on last row in excel

    """
    file_storage.stream.seek(0)

    # Load the workbook directly from the file stream
    try:
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        return f"Error loading workbook: {e}"

    sheet = wb.active  # Or wb['Sheet1']
    # list of all segments
    # Track two separate subtotals
    summaries_col_15 = []
    summaries_col_17 = []

    subtotal_15 = 0
    subtotal_17 = 0
    # Assuming header is row 1, data starts at row 2
    starting_row = 2
    # Column O in excel is the 15th column
    min_column_index = 15
    max_column_index = 17

    for row in sheet.iter_rows(
        min_row=starting_row, max_col=max_column_index, min_col=min_column_index
    ):
        cell_15 = row[0]
        cell_17 = row[2]

        border = cell_15.border  # Assuming borders are consistent across the row
        val_15 = cell_15.value if isinstance(cell_15.value, (int, float)) else 0
        val_17 = cell_17.value if isinstance(cell_17.value, (int, float)) else 0

        # 1. STOP & CAPTURE: Grand Total Detection
        if border.top.style == "medium" and border.bottom.style == "medium":
            grand_total_from_file_15 = val_15
            grand_total_from_file_17 = val_17
            print(f"--- VALIDATION: Grand Total Found: {grand_total_from_file_15} ---")
            print(f"--- VALIDATION: Grand Total Found: {grand_total_from_file_17} ---")
            break

        # 2. TRIGGER TOP BORDER: New Group Starts (Top Style)
        if border.top.style == "medium":
            if subtotal_15 > 0 or subtotal_17 > 0:
                summaries_col_15.append(subtotal_15)
                summaries_col_17.append(subtotal_17)
            subtotal_15 = val_15
            subtotal_17 = val_17

        # 3. TRIGGER BOTTOM BORDER: Current Group Ends (Bottom Style)
        elif border.bottom.style == "medium":
            subtotal_15 += val_15
            subtotal_17 += val_17
            if subtotal_15 > 0 or subtotal_17 > 0:
                summaries_col_15.append(subtotal_15)
                summaries_col_17.append(subtotal_17)
            subtotal_15 = 0
            subtotal_17 = 0

        # 4. NORMAL: Regular accumulation
        else:
            subtotal_15 += val_15
            subtotal_17 += val_17

    # Final cleanup if data remains (in case Grand Total row was missing)
    if subtotal_15 > 0 or subtotal_17 > 0:
        summaries_col_15.append(subtotal_15)
        summaries_col_17.append(subtotal_17)

    # --- Verification Logic ---
    calculated_total_15 = sum(summaries_col_15)
    is_valid_15 = calculated_total_15 == grand_total_from_file_15

    calculated_total_17 = sum(summaries_col_17)
    is_valid_17 = calculated_total_17 == grand_total_from_file_17

    return {
        "gpon": {
            "segments": summaries_col_15,
            "calculated_total": calculated_total_15,
            "excel_grand_total": grand_total_from_file_15,
            "match": is_valid_15,
        },
        "xdsl": {
            "segments": summaries_col_17,
            "calculated_total": calculated_total_17,
            "excel_grand_total": grand_total_from_file_17,
            "match": is_valid_17,
        },
    }


def save_imported_segments_to_db(payload, target_date=None):
    """
    Map the Excel segment indexses (0, 1, 2...) to your DB city_id
    """

    if not ftth_view_required():
        return False, "Niste autorizovani."

    # Backend sanitization
    # If target_date is provided than it is from (Admin) dashboard, use it.
    # Otherwise (User) it is from regular user domain, get current month end.
    if target_date:
        # If target_date is a string from JSON, convert to object
        if isinstance(target_date, str):
            dt = datetime.strptime(target_date, "%Y-%m-%d")
        else:
            dt = target_date

        # FORCE the date to the last day of that specific month
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        month_to_save = dt.replace(day=last_day).date()

    else:
        # Regular user logic: get actual current month end
        month_to_save = get_previous_month_end()

    # Mapping: Payload Key -> access_type_id
    type_mapping = {"gpon": 1, "xdsl": 2}

    # SEG 1 BANJA LUKA -> (city_id=3)
    # SEG 2 BIJELJINA-> (city_id=6)
    # SEG 3 BRCKO-> (city_id=7)
    # SEG 4 DOBOJ-> (city_id=5)
    # SEG 5 SARAJEVO-> (city_id=9)
    # SEG 6 PRIJEDOR-> (city_id=4)
    # SEG 7 TREBINJE-> (city_id=11)
    # SEG 8 FOCA-> (city_id=10)
    # SEG 9 ZVORNIK-> (city_id=8)
    # Map the Excel segment index (0, 1, 2...) to your DB city_id
    # Ensure this order matches your Excel segments exactly!
    city_mapping = [3, 6, 7, 5, 9, 4, 11, 10, 8]
    total_inserted = 0

    try:
        # Loop through each access type (gpon, then xdsl)
        for type_key, segments in payload.items():
            access_type_id = type_mapping.get(type_key)

            if not access_type_id:
                continue

            for index, value in enumerate(segments):
                # Guard against Excel providing more segments than we have city mappings
                if index >= len(city_mapping):
                    break

                city_id = city_mapping[index]
                quantity = int(value or 0)

                # UPSERT
                db.session.execute(
                    text("""
                        INSERT INTO access_inventory (city_id, access_type_id, month_end, quantity)
                        VALUES (:city_id,:at_id, :month_end, :quantity)
                        ON CONFLICT (city_id, access_type_id, month_end)
                        DO UPDATE SET quantity = EXCLUDED.quantity, updated_at = NOW();
                    """),
                    {
                        "city_id": city_id,
                        "at_id": access_type_id,
                        "month_end": month_to_save,
                        "quantity": quantity,
                    },
                )
                total_inserted += 1

        db.session.commit()
        return (
            True,
            f"Uspješno ažurirano {total_inserted} zapisa za {month_to_save}.",
        )

    except Exception as e:
        db.session.rollback()
        print(e)
        return False, "Greška prilikom čuvanja podataka."


# -------------------------
# INTERNAL HELPERS
# -------------------------
def _group_records(records, month_keys):
    grouped = {}

    for row in records:
        cityid = row["id"] or TOTAL_KEY

        if cityid not in grouped:
            grouped[cityid] = {
                "id": cityid,
                "name": row["name"],
                "last_updated": row["last_updated"],
                "is_total": cityid == TOTAL_KEY,
                "dates": {month: {"quantity": 0} for month in month_keys},
            }

        for month in month_keys:
            grouped[cityid]["dates"][month]["quantity"] = row.get(month, 0)

    return list(grouped.values())
