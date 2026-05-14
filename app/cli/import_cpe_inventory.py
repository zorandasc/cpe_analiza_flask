import click
from flask.cli import with_appcontext
from datetime import datetime, time
import re
from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from app.models import Cities, CpeInventory, CpeTypes
from app.extensions import db


# flask import_cpe_inventory xxx.xlsx > output.log 2>&1
@click.command("import_cpe_inventory")
@click.argument("excel_path")
@with_appcontext
def import_cpe_inventory_command(excel_path):
    with open(excel_path, "rb") as f:
        records = import_cpe_inventory_from_excel(f)

        bulk_upsert_cpe_dismantle(records)


def import_cpe_inventory_from_excel(file_stream):
    """
    1.Load the workbook
    2.Iterate through sheets
    3.Find tables
    4.Parse table, header, rows
    5.Append normalized dicts first.
    6.validate
    7.deduplicate
    8.bulk upsert
    """

    # 1. Load the workbook directly from the file stream
    try:
        workbook = load_workbook(file_stream, data_only=True)
    except Exception as e:
        return f"Error loading workbook: {e}"

    # Imena i id gradova iz baze
    CITY_MAP = {normalize(city.name): city.id for city in Cities.query.all()}

    # DINAMICKO PAMIRANJE IMENA EXCELA I ID U BAZI
    CPE_MAP = {normalize(cpe.label): cpe.id for cpe in CpeTypes.query.all()}

    # + NEKA SPECIFICNA IMEAN STATICKI MAPIRANA
    CPE_MAP[normalize("IAD-a H267N /  HG658V2 / Zyxel")] = 1
    CPE_MAP[normalize("IAD-a H267N /  HG658V2 / Zyxel / Skyworth")] = 1
    CPE_MAP[normalize("STB-ova Arris VIP4205/ VIP4302/1113")] = 2
    CPE_MAP[normalize("STB-ova Arris VIP4205/ VIP4302")] = 2
    CPE_MAP[normalize("STB-ova Arris VIP5305")] = 3
    CPE_MAP[normalize("STB-ova EKT DIN4805V 4K")] = 4
    CPE_MAP[normalize("STB-ova EKT DIN7005V HD")] = 5
    CPE_MAP[normalize("Skyworth STB HD/4K HP44H")] = 6

    records = []

    print(f"Workbook sheets: {len(workbook.worksheets)}")

    # 2. Iterate through sheets
    for sheet in workbook.worksheets:
        # Find week_end from sheet title
        sheet_title = sheet.title.strip().rstrip(".")

        print(f"Processing: {sheet.title}")

        week_end = datetime.strptime(sheet_title, "%d.%m.%Y").date()

        # Inside each sheet Find table "Stanje"
        for row in sheet.iter_rows():
            for cell in row:
                # FIND START OF COMPLETET TABLE
                if cell.value == "Stanje":
                    data_start_row = cell.row + 3  # ROW OF FIRST CITY
                    start_col = 3  # COLUMN OF FIRST CPE (C column in excel)
                else:
                    continue

                # for every sheet build dinamicaly MAPPING BETWEEN
                # excel_colums:cpe_type_id}
                column_to_cpe_id = dynamic_build_cpe_id(
                    CPE_MAP, sheet, cell.row, start_col
                )

                print("column_to_cpe_id", column_to_cpe_id)

                # Start row row that hlds city + data
                row_num = data_start_row

                while True:
                    # Get city name from excel
                    city_name = sheet.cell(row=row_num, column=1).value

                    normalize_city = normalize(city_name)

                    if normalize_city == "ukupno":
                        row_num += 1
                        continue

                    # Mapiranje izmedju grada excela i baze
                    city_id = CITY_MAP.get(normalize_city)

                    # table ended EXIT WHILE LOOP
                    if city_id is None:
                        print(f"End of table at row {row_num}: {city_name}")

                        break

                    for col_num, cpe_type_id in sorted(column_to_cpe_id.items()):
                        raw_cell = sheet.cell(row=row_num, column=col_num).value

                        parsed_quantity = parse_excel_cell(raw_cell)

                        if parsed_quantity is None:
                            print(
                                f"Invalid quantity "
                                f"Sheet={sheet.title}"
                                f"at row={row_num}, "
                                f"col={col_num}: "
                                f"{raw_cell}"
                            )
                            continue

                        row_data = {
                            "city_id": city_id,
                            "cpe_type_id": cpe_type_id,
                            "quantity": parsed_quantity,
                            "week_end": week_end,
                            "created_at": datetime.combine(week_end, time.min),
                            "updated_at": datetime.combine(week_end, time.min),
                        }

                        records.append(row_data)

                    # NEXT ROW
                    row_num += 1

    return records


def dynamic_build_cpe_id(cpe_map, sheet, header_row, start_col):

    column_to_cpe_id = {}  # Holds dynamic mappings of excel_colums:cpe_type_id

    empty_streak = 0  # counter of empty columns
    max_empty_streak = 3  # stop after N consecutive empty columns

    col_num = start_col

    # In header_row start parsing column
    while True:
        # Get name of cpe from columns in headers
        cpe_name = sheet.cell(row=header_row, column=col_num).value

        normalized = normalize(cpe_name)

        if not normalized:
            empty_streak += 1

            # assume table ended
            if empty_streak >= max_empty_streak:
                break

            col_num += 1
            continue

        # reset empty streak counter
        empty_streak = 0

        # Map that name to id from db
        cpe_type_id = cpe_map.get(normalized)

        # No id in db
        if cpe_type_id is None:
            print(f"Unknown CPE header: {cpe_name}")
            col_num += 1
            continue

        # if id in db found, propagate merged header
        column_to_cpe_id[col_num] = cpe_type_id

        col_num += 1

    return column_to_cpe_id


def normalize(value):

    if not value:
        return ""

    value = str(value)

    value = value.replace("\n", " ")
    value = value.replace("\xa0", " ")

    value = value.strip().lower()

    value = re.sub(r"\s+", " ", value)

    return value


def parse_excel_cell(raw_cell):
    if raw_cell is None:
        return 0

    # Excel sometimes returns floats
    if isinstance(raw_cell, (int, float)):
        return int(raw_cell)

    # Convert everything else to string
    value = str(raw_cell).strip()

    if value in ["", "-", "/", "*"]:
        return 0

    # Normal integer string
    try:
        return int(float(value))
    except Exception:
        return None


def bulk_upsert_cpe_dismantle(records):
    if not records:
        print("No records to import.")
        return

    print(f"Prepared {len(records)} records")
    try:
        stmt = insert(CpeInventory).values(records)

        stmt = stmt.on_conflict_do_update(
            constraint="uq_city_cpe_week",
            set_={
                "quantity": stmt.excluded.quantity,
                "updated_at": db.func.now(),
            },
        )

        print("Executing bulk upsert...")

        db.session.execute(stmt)

        db.session.commit()

        print(f"Upserted {len(records)} records.")

    except Exception:
        db.session.rollback()
        raise
