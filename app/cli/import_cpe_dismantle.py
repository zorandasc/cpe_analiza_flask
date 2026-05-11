import click
from flask.cli import with_appcontext
from datetime import datetime
import re
from collections import defaultdict
from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from app.models import CpeDismantle
from app.extensions import db

# MAPIRANJE REDNOG BROJA EXCEL KOLONE U ID TOG CPE TIPA U POSTGRES DB
COLUMN_TO_CPE_ID = {
    # exce_first_column_number: cpe_type_id
    3: 1,
    5: 2,
    7: 3,
    9: 7,
    11: 8,
    12: 4,
    13: 5,
}

# MAPIRANJE POZICIJE GRADA U EXCEL TABELI I ID TOG GRADA U POSTGRES DB
ROW_TO_CITY_ID = {
    # exce_relativ_table_row_number: city_id
    1: 3,
    2: 4,
    3: 5,
    4: 6,
    5: 8,
    6: 9,
    7: 10,
    8: 11,
}


# flask import_cpe_dismantle xxx.xlsx > output.log 2>&1
@click.command("import_cpe_dismantle")
@click.argument("excel_path")
@with_appcontext
def import_cpe_dismantle_command(excel_path):
    with open(excel_path, "rb") as f:
        records = import_cpe_dismantle_from_excel(f)

        bulk_upsert_cpe_dismantle(records)


def import_cpe_dismantle_from_excel(file_stream):
    """
    1.Load the workbook
    2.Iterate through sheets
    3.Find tables
    4.Parse table, header, rows
    5.Append normalized dicts first.
    6.validate
    7.deduplicate
    8.bulk upsert
    """

    # 1. Load the workbook directly from the file stream
    try:
        workbook = load_workbook(file_stream, data_only=True)
    except Exception as e:
        return f"Error loading workbook: {e}"

    records = []

    print(f"Workbook sheets: {len(workbook.worksheets)}")

    # 2. Iterate through sheets
    for sheet in workbook.worksheets:
        # Find week_end from sheet title
        sheet_title = sheet.title.strip().rstrip(".")

        print(f"Processing: {sheet.title}")

        week_end = datetime.strptime(sheet_title, "%d.%m.%Y").date()

        # Inside each sheet Find table dimsantle_complete and dismantle_missing
        for row in sheet.iter_rows():
            for cell in row:
                # FIND START OF COMPLETET TABLE
                if cell.value == "Stanje demontirane ispravne i kompletne TO":
                    table_type = "COMP"
                    # ROW OF FIRST CITY
                    data_start_row = cell.row + 3
                elif cell.value == "Stanje demontirane ispravne i nekompletne TO":
                    table_type = "MISS"
                    # ROW OF FIRST CITY
                    data_start_row = cell.row + 3
                else:
                    continue

                # ROW OF LAST CITY
                data_end_row = data_start_row + len(ROW_TO_CITY_ID) - 1

                # COLUMN OF FIRST CPE (C column in excel)
                start_col = 3

                # COLUMN OF LAST CPE  (M column in excel)
                stop_col = 13

                for row_num in range(data_start_row, data_end_row + 1):
                    # in first iteration row_num=data_start_row
                    # This determines relative index in ROW_TO_CITY_ID from absolute excel row
                    # Excel row     Relative row
                    # --------------------------------
                    # 12            1
                    # 13            2
                    # 14            3
                    # 15            4
                    relative_row = row_num - data_start_row + 1

                    city_id = ROW_TO_CITY_ID.get(relative_row)

                    if not city_id:
                        print(f"Unknown relative row: {relative_row}")
                        continue

                    for col_num in range(start_col, stop_col + 1):
                        cpe_type_id = COLUMN_TO_CPE_ID.get(col_num)

                        if not cpe_type_id:
                            continue

                        raw_cell = sheet.cell(row=row_num, column=col_num).value

                        # choose parser function
                        parser = TABLE_PARSERS[table_type]

                        parsed_rows = parser(raw_cell, cpe_type_id)

                        if parsed_rows is None:
                            print(
                                f"Invalid quantity "
                                f"Sheet={sheet.title}"
                                f"at row={row_num}, "
                                f"col={col_num}: "
                                f"{raw_cell}"
                            )
                            continue

                        for parsed in parsed_rows:
                            row_data = {
                                "city_id": city_id,
                                "cpe_type_id": cpe_type_id,
                                "dismantle_type_id": parsed["dismantle_type_id"],
                                "quantity": parsed["quantity"],
                                "week_end": week_end,
                            }

                            records.append(row_data)

    return records


# parse_quantity(5) 5
# parse_quantity("10") 10
# parse_quantity("10/20") 30
# parse_quantity("10 / 20 / 5") 35
# parse_quantity("-") 0
# parse_quantity(None) 0
def parse_complete_quantity(raw_cell, _):
    if raw_cell is None:
        return [{"quantity": 0, "dismantle_type_id": 1}]

    # Excel sometimes returns floats
    if isinstance(raw_cell, (int, float)):
        return [{"quantity": int(raw_cell), "dismantle_type_id": 1}]

    # Convert everything else to string
    value = str(raw_cell).strip()

    if value in ["", "-", "/", "*"]:
        return [{"quantity": 0, "dismantle_type_id": 1}]

    # Handle values like: 10/20
    if "/" in value:
        try:
            parts = value.split("/")
            total = sum(int(float(part.strip())) for part in parts if part.strip())
            return [{"quantity": total, "dismantle_type_id": 1}]
        except Exception:
            return None

    # Normal integer string
    try:
        return [{"quantity": int(float(value)), "dismantle_type_id": 1}]
    except Exception:
        return None


# Posible cell values:
# number CODE =>[{"quantity":number, "dismantle_type_id"=DISMANTLE_CODE_MAP(CODE)}]
# number => [{"quantity":number, "dismantle_type_id"= DEFAULT_DISMANTLE_BY_CPE[cpe_type_id])}]
# number1 CODE1/number2 CODE1 => [{"quantity":number1+number2, "dismantle_type_id"=DISMANTLE_CODE_MAP(CODE1)}]
# number1 CODE1/number2 CODE2 => [{"quantity":number1, "dismantle_type_id"=DISMANTLE_CODE_MAP(CODE1)},
# {"quantity":number2, "dismantle_type_id"=DISMANTLE_CODE_MAP(CODE2)}]
# number1/number2   =>[{"quantity":number1+number2, "dismantle_type_id"= DEFAULT_DISMANTLE_BY_CPE[cpe_type_id])}]
def parse_missing_quantity(raw_cell, cpe_type_id):

    # PHASE 1 — syntax parsing
    # Convert string from one excel cell into normalized tokens.
    # excel cell = 100 ND / 200 NA
    # tokens=[(100, "ND"), (200, "NDIA"), (100, None),...]
    tokens = parse_tokens(raw_cell)

    # SUME QUANTITY OF SAME DISMANTLE CODE: 100 ND/200 ND= 300 ND
    aggregated = defaultdict(int)

    # PHASE 2 — semantic interpretation
    for token in tokens:
        # Interpret: missing code, dismantle mapping
        dismantle_type_id = resolve_dismantle_type(cpe_type_id, token["code"])

        # aggregation of same codes
        aggregated[dismantle_type_id] += token["quantity"]

    return [
        {
            "quantity": quantity,
            "dismantle_type_id": dismantle_type_id,
        }
        for dismantle_type_id, quantity in aggregated.items()
    ]


TOKEN_REGEX = re.compile(r"(\d+)\s*([A-Z]+)?")


# Convert ONE EXCEL CELL string into one or many normalized tokens.
def parse_tokens(value):

    if value is None:
        return []

    # normalize
    value = str(value).strip().upper()

    if value in ["", "-", "/", "*"]:
        return []

    result = []

    # split by "/"
    parts = value.split("/")

    for part in parts:
        part = part.strip()

        # (100, "ND"),
        match = TOKEN_REGEX.fullmatch(part)

        if not match:
            raise ValueError(f"Invalid token: {part}")

        quantity = int(match.group(1))

        code = match.group(2)

        result.append({"quantity": quantity, "code": code})

    return result


DISMANTLE_CODE_MAP = {
    "ND": 2,
    "NA": 3,
    "NDIA": 4,
}

# FOR excel CELL THAT HAVE ONLY NUMBER WE CHOOSE
# DEFAULT DISMANTLE_ID IN REGARD OF CPE_TYPE_ID
DEFAULT_DISMANTLE_BY_CPE = {
    # cpe_type_id: dismantle_type_id
    1: 3,  # IAD (cpe_type_id=1) -> NA
    2: 2,  # VIP4205 /VIP4302 (cpe_type_id=2) -> ND
    3: 2,  # Arris VIP5305 (cpe_type_id=3) -> ND
    7: 3,  # ONT (HUAWEI) (cpe_type_id=3) -> NA
    8: 3,  # ONT (NOKIA) (cpe_type_id=8) -> NA
    4: 2,  # EKT DIN4805V (cpe_type_id=4) -> ND
    5: 2,  # EKT DIN7005V HD (cpe_type_id=5) -> ND
}


# RESOLVE DISMANTLE CODE from tokens TO ID
def resolve_dismantle_type(cpe_type_id, dismantle_code):
    if dismantle_code:
        dismantle_code = dismantle_code.strip().upper()

        # IF THERE IS DISMANTLE CODE WE USE ID OF THAT CODE
        return DISMANTLE_CODE_MAP.get(dismantle_code)

    # IF THERE IS NO DISMANTLE CODE WE USE DEFAULT FOR THAT CPE_TYPE
    return DEFAULT_DISMANTLE_BY_CPE.get(cpe_type_id)


TABLE_PARSERS = {
    "COMP": parse_complete_quantity,
    "MISS": parse_missing_quantity,
}


def bulk_upsert_cpe_dismantle(records):
    if not records:
        print("No records to import.")
        return

    print(f"Prepared {len(records)} records")
    try:
        stmt = insert(CpeDismantle).values(records)

        stmt = stmt.on_conflict_do_update(
            constraint="uq_city_cpe_dismantle_week",
            set_={
                "quantity": stmt.excluded.quantity,
                "updated_at": db.func.now(),
            },
        )

        print("Executing bulk upsert...")

        db.session.execute(stmt)

        db.session.commit()

        print(f"Upserted {len(records)} records.")

    except Exception:
        db.session.rollback()
        raise
