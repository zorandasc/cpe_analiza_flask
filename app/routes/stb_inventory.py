from datetime import datetime
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from app.services.stb_inventory import (
    get_stb_records_view_data,
    update_recent_stb_inventory,
    update_iptv_users_count,
    get_stb_records_excel_export,
)

stb_inventory_bp = Blueprint(
    "stb_inventory",
    __name__,
    url_prefix="/stb-records",
)


@stb_inventory_bp.route("/")
@login_required
def stb_records():
    data = get_stb_records_view_data()
    return render_template("stb_records.html", **data)


@stb_inventory_bp.route("/update_stb", methods=["POST"])
@login_required
def update_stb_inventory():
    success, message = update_recent_stb_inventory(request.form)
    flash(message, "success" if success else "danger")
    return redirect(url_for("stb_inventory.stb_records"))


@stb_inventory_bp.route("/update_iptv_users", methods=["POST"])
@login_required
def update_iptv_users():
    success, message = update_iptv_users_count(request.form)
    flash(message, "success" if success else "danger")
    return redirect(url_for("stb_inventory.stb_records"))


@stb_inventory_bp.route("/export/stb-records.xlsx")
@login_required
def export_stb_records_excel():
    # service
    headers_stb, rows_stb, headers_iptv, rows_iptv, current_week_end = (
        get_stb_records_excel_export()
    )

    wb = Workbook()

    # ---------------------------
    # STB INVENTORY SHEET
    # ---------------------------
    ws_stb = wb.active
    ws_stb.title = "STB oprema"

    ws_stb.append(headers_stb)
    for row in rows_stb:
        ws_stb.append(row)

    # styling
    # bold header
    for cell in ws_stb[1]:
        cell.font = Font(bold=True)

    ws_stb.freeze_panes = "A2"

    # auto-width columns (nice UX)
    for col in ws_stb.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                # odredi max length
                max_length = max(max_length, len(str(cell.value)))
        ws_stb.column_dimensions[col[0].column_letter].width = max_length + 2

    # ---------------------------
    # IPTV USERS SHEET
    # ---------------------------
    ws_iptv = wb.create_sheet("IPTV korisnici")

    ws_iptv.append(headers_iptv)
    for row in rows_iptv:
        ws_iptv.append(row)

    # sty;yng bold header
    for cell in ws_iptv[1]:
        cell.font = Font(bold=True)

    ws_iptv.freeze_panes = "B2"

    # increse column width to be max of content
    for col in ws_iptv.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_iptv.column_dimensions[col[0].column_letter].width = max_len + 2

    # ---------------------------
    # INFO SHEET
    # ---------------------------
    meta = wb.create_sheet("Info")
    meta.append(["Kreirano:", datetime.now().strftime("%d-%m-%Y %H:%M")])
    meta.append(["Sedmica ažuriranja:", current_week_end.strftime("%d-%m-%Y")])

    # ---------------------------
    # EXPORT
    # ---------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="stanje_stb_opreme_iptv_korisnika.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
