from datetime import datetime
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    jsonify,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from app.services.ont_inventory import (
    get_ont_records_view_data,
    update_recent_ont_inventory,
    get_ont_records_excel_export,
    parce_excel_segments,
    save_imported_segments_to_db,
)

ont_inventory_bp = Blueprint(
    "ont_inventory",
    __name__,
    url_prefix="/ont-records",
)


@ont_inventory_bp.route("/")
@login_required
def ont_records():
    data = get_ont_records_view_data()
    return render_template("ont_records.html", **data)


@ont_inventory_bp.route("/update_ont", methods=["POST"])
@login_required  # AUTENTIFICATION
def update_ont_inventory():
    success, message = update_recent_ont_inventory(request.form)
    flash(message, "success" if success else "danger")
    return redirect(url_for("ont_inventory.ont_records"))


@ont_inventory_bp.route("/export/ont-records.xlsx")
@login_required
def export_ont_records_excel():
    headers, rows, current_month_end = get_ont_records_excel_export()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stanje ONT Opreme"

    meta = wb.create_sheet("Info")
    meta.append(["Kreirano:", datetime.now().strftime("%d-%m-%Y %H:%M")])
    meta.append(["Mjesec Ažuriranja:", current_month_end.strftime("%d-%m-%Y %H:%M")])

    ws.append(headers)

    for row in rows:
        ws.append(row)

    # auto-width columns (nice UX)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="stanje_ont_opreme.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# called from js inside ont_records.html
@ont_inventory_bp.route("/upload-excel", methods=["POST"])
@login_required
def import_ont_records_excel():
    if "file" not in request.files:
        return "No file part", 400

    file = request.files["file"]

    # The dictionary returned here contains 'segments', 'match', etc.
    results = parce_excel_segments(file)

    # RETRUN PARSED SEGMENTS TO MODAL FOR DISPLAY
    return results  # Flask converts dict to JSON automatically


@ont_inventory_bp.route("/save-segments", methods=["POST"])
@login_required
def save_imported_segments():
    data = request.get_json()

    segments = data.get("segments", [])

    if not segments:
        return jsonify(
            {
                "success": False,
                "message": "error: Nema podataka",
            }
        ), 400

    success, message = save_imported_segments_to_db(segments,[])

    flash(message, "success" if success else "danger")

    return jsonify(
        {
            "success": success,
            "message": message,
        }
    ), 200 if success else 403
