from datetime import datetime
from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    flash,
    url_for,
    send_file,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.services.cpe_broken import get_cpe_broken_view_data
from app.services.cpe_dismantle import (
    get_cpe_dismantle_view_data,
    update_cpe_dismantle,
    get_cpe_dismantle_history,
    get_cpe_dismantle_excel_export,
    get_missing_subcolumns,
)


cpe_dismantle_bp = Blueprint(
    "cpe_dismantle_inventory",
    __name__,
    url_prefix="/cpe-dismantle-records",
)


@cpe_dismantle_bp.route("/")
@login_required
def cpe_dismantle_records():
    data_dismantle = get_cpe_dismantle_view_data()

    data_broken = get_cpe_broken_view_data()

    return render_template(
        "cpe_dismantle.html",
        data_broken=data_broken,
        **data_dismantle
    )


@cpe_dismantle_bp.route("/update", methods=["POST"])
@login_required
def cpe_dismantle_update():
    data = request.get_json(silent=True)

    success, message = update_cpe_dismantle(data or {})

    flash(message, "success" if success else "danger")

    return jsonify(
        {
            "success": success,
            "message": message,
        }
    ), 200 if success else 403


@cpe_dismantle_bp.route("/history/<int:id>/<category>")
@login_required
def cpe_dismantle_city_history(id, category):
    page = request.args.get("page", 1, int)

    per_page = 20

    city, records, schema_list, category, error = get_cpe_dismantle_history(
        id, page, per_page, category
    )

    if error:
        flash(error, "danger")
        return redirect(url_for("main.home"))

    return render_template(
        "cpe_dismantle_history.html",
        records=records,
        category=category,
        schema=schema_list,
        city=city,
    )


@cpe_dismantle_bp.route("/export/cpe-dismantle.xlsx/<category>")
@login_required
def export_cpe_dismantle_excel(category):
    headers_main, headers_sub, rows, schema_list, current_week_end = (
        get_cpe_dismantle_excel_export(mode=category)
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "CPE - kompletno" if category == "complete" else "CPE - nekompletno"

    meta = wb.create_sheet("Info")
    meta.append(["Kreirano:", datetime.now().strftime("%d-%m-%Y %H:%M")])
    meta.append(["Sedmica Ažuriranja:", current_week_end.strftime("%d-%m-%Y %H:%M")])

    # headers
    ws.append(headers_main)
    ws.append(headers_sub)

    # merge grouped headers
    if category != "complete":
        col = 2  # City is column 1
        for cpe in schema_list:
            subcols = get_missing_subcolumns(cpe)

            if not subcols:
                continue

            start_col = col
            end_col = col + len(subcols) - 1

            # openpyxl creates MergedCell objects.
            ws.merge_cells(
                start_row=1,
                start_column=start_col,
                end_row=1,
                end_column=end_col,
            )
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")
            col = end_col + 1

    # rows
    for row in rows:
        ws.append(row)

    # formatting
    # bold both header rows:
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)

    # two header rows
    ws.freeze_panes = "A3"

    # Correct way to auto-size columns when merged cells exist
    # automatically adjusting column widths to fit their content.
    # enumerate(..., start=1)gives us the index i (1, 2, 3...) which corresponds to Excel's column numbering.
    for i, col in enumerate(ws.columns, start=1):
        # tracker for the longest piece of text found in the current column.
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # get_column_letter(i) converts the number (like 1) into an Excel letter (like A
        # It sets the column width to the longest string length found, plus 2 units of
        # padding so the text doesn't touch the cell borders.
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        "Demontirana_kompletna_cpe_oprema.xlsx"
        if category == "complete"
        else "Demontirana_nekompletna_cpe_oprema.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
