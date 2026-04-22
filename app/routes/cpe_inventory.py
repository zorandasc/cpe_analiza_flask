from datetime import datetime
from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    flash,
    url_for,
    send_file,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from app.services.cpe_inventory import (
    get_cpe_records_view_data,
    get_cpe_records_history,
    update_cpe_records,
    get_cpe_records_excel_export,
    get_cpe_records_subcities,
)


cpe_inventory_bp = Blueprint(
    "cpe_inventory",
    __name__,
    url_prefix="/cpe-records",
)


@cpe_inventory_bp.route("/")
@login_required
def cpe_records():
    data = get_cpe_records_view_data()

    return render_template("cpe_records.html", **data)


@cpe_inventory_bp.route("/subcities/<int:id>")
@login_required
def cpe_records_subcities(id):

    data = get_cpe_records_subcities(city_id=id)

    return render_template("cpe_records_subcities.html", **data)


# UPDATE ROUTE FOR CPE-RECORDS TABLE, CALLED FROM INSIDE FORME INSIDE cpe-record
# THIS IS JSON API ROUTE, IT DOESNOT RETURN HTNL PAGE
@cpe_inventory_bp.route("/update", methods=["POST"])
@login_required
def cpe_records_update():
    data = request.get_json(silent=True)

    success, message = update_cpe_records(data or {})

    flash(message, "success" if success else "danger")

    # return {"status": "ok"}
    return jsonify(
        {
            "success": success,
            "message": message,
        }
    ), 200 if success else 403


@cpe_inventory_bp.route("/history/<int:id>")
@login_required
def cpe_records_city_history(id):
    page = request.args.get("page", 1, int)

    per_page = 20

    scope = request.args.get("scope", "city")

    city, records, schema_list, error = get_cpe_records_history(id, page, per_page, scope)
   
    if error:
        flash(error, "danger")
        return redirect(url_for("main.home"))
    
    if scope == 'major':
        back_url = url_for('cpe_inventory.cpe_records')
    elif city.parent_city_id:
        back_url = url_for('cpe_inventory.cpe_records_subcities', id=city.parent_city_id)
    else:
        back_url = url_for('cpe_inventory.cpe_records_subcities', id=city.id)

    return render_template(
        "cpe_records_history.html",
        records=records,
        schema=schema_list,
        city=city,
        back_url=back_url
    )


@cpe_inventory_bp.route("/export/cpe-records.xlsx")
@login_required
def export_cpe_records_excel():
    warehouse_fill = PatternFill("solid", fgColor="EEEEEE")
    warehouse_font = Font(color="666666")
    total_font = Font(bold=True)

    headers, rows, current_week_end = get_cpe_records_excel_export()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stanje CPE Opreme"

    meta = wb.create_sheet("Info")
    meta.append(["Kreirano:", datetime.now().strftime("%d-%m-%Y %H:%M")])
    meta.append(["Sedmica Ažuriranja:", current_week_end.strftime("%d-%m-%Y %H:%M")])

    # You must insert newline characters (\n) into header text
    # to wrap text in headaer of excel
    formatted_headers = [h.replace(" ", "\n/ ") for h in headers]
    ws.append(formatted_headers)

    for row_data in rows:
        ws.append(row_data["values"])
        # ukupan broj celija u rowu
        excel_row = ws.max_row

        # red UKUPNO bolduj
        if row_data["city_id"] is None:
            # za sve celije u rowu
            for cell in ws[excel_row]:
                cell.font = total_font

        # red Raspoloziva oprema oboji u sivo
        elif row_data["city_id"] == 13:
            # za sve celije u rowu
            for cell in ws[excel_row]:
                cell.fill = warehouse_fill
                cell.font = warehouse_font

    # style header wrap text + bold text
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )

    # Excel will only wrap text if row height allows it.
    # set header row height manually:
    ws.row_dimensions[1].height = 65

    # smart auto-width
    # When calculating column width:
    # only consider the longest line, not full string length.
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_length = max(max_length, max(len(li) for li in lines))

        ws.column_dimensions[col_letter].width = max_length + 2

    ws.freeze_panes = "A2"

    output = BytesIO()

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="stanje_cpe_opreme.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
